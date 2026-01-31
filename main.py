import tkinter as tk

import search_data


from tkinter import *
from tkinter import font, messagebox
from tkinter import ttk

from openpyxl import load_workbook


# ================== DATES ==============================================

months = [" ", "січень", "лютий", "березень", "квітень", "травень", "червень", "липень", "серпень", "вересень",
                  "жовтень", "листопад", "грудень"]

energy = ['газу', 'електроенергії', 'води']

entries = []
address = [' ', 'КВІТНЕВА-8', 'КВІТНЕВА-30-5']

text = ''
text2 = ''
year = 2026
table_name = f'year{year}'
# ================================== FUNCTIONS   ============================================

def selected(event):
    global text, text2
    widget = event.widget
    tag = getattr(widget, "tag", None)
    text1 = widget.get()



    if tag == "month":
        text = text1
        label_report(lf_H1, 2, 0, text)

    if tag == "address":
        text2 = text1
        label_report(lf_H1, 2, 1, text2)




def add_date():

    match text2:
        case '':
            messagebox.showerror("Помилка вводу", "Оберите адресу !!!")

        case "КВІТНЕВА-8":                   # Замінити параметри

            search_data.find_in_excel_table('K_8.xlsx', 'A1', f'{table_name}', 'Дата', f'{text}')

            ''' wb = load_workbook("K_8.xlsx")
                ws = wb.active
                data_gas_num = entries[0].get()
                data_light_num = entries[1].get()
                data_water_num = entries[2].get()

                ws['C21'] = data_gas_num
                ws['J21'] = data_light_num
                ws['P21'] = data_water_num
                wb.save("K_8.xlsx")
            '''



        case "КВІТНЕВА-30-5":               # Замінити параметри

            search_data.find_in_excel_table('k_30.xlsx', 'A1', f'{table_name}', 'Дата', f'{text}')

            '''wb = load_workbook("k_30.xlsx")
            ws = wb.active
            data_gas_num = entries[0].get()
            data_light_num = entries[1].get()
            data_water_num = entries[2].get()

            ws['C21'] = data_gas_num
            ws['J21'] = data_light_num
            ws['P21'] = data_water_num
            wb.save("k_30.xlsx")
            '''
    messagebox.showinfo("Інформація", "Данні додані до файлу")







def label_report(frame, row, column, value):

    report_label = ttk.Label(frame, text=f'Ви вибрали:{value}', font=courier_14, foreground='green')
    report_label.grid(row=row, column=column, ipadx=6, ipady=6, padx=55, pady=5)
    print('l>>>>>')

def validate_input(new_value):
    if new_value == "":  # дозволяємо очищення поля
        return True
    if not new_value.isdigit():
        messagebox.showerror("Помилка вводу", "Дозволені тільки цифри!")
        return False
    if len(new_value) > 5:
        messagebox.showerror("Помилка вводу", "Максимум 5 цифр!")
        return False
    return True


# =/////////////////===================  MAIN ==========================////////////////////////////////

root = tk.Tk()
root.title("MAIN WINDOW")
root.geometry("900x800")
root.resizable(False, False)

courier_10 = font.Font(family="Courier", size=10, weight=font.BOLD)
courier_14 = font.Font(family="Courier", size=14, weight=font.BOLD)
courier_18 = font.Font(family="Courier", size=18, weight=font.BOLD)
width_frame = 800

label = tk.Label(root, text="ДОДАВАННЯ ПОКАЗНИКІВ ЛІЧИЛЬНИКІВ У ФАЙЛ", fg="BLUE", font=courier_18)
label.grid(row=0, column=0, columnspan=3, ipadx=6, ipady=6, padx=5, pady=15)

# ========================  main frame  =======================================

lf_MF = ttk.Frame(root, borderwidth=10, relief=SUNKEN)
lf_MF.config(width=850, height=600)


lf_H1 = ttk.Frame(lf_MF, borderwidth=10, relief=SUNKEN)
lf_H1.config(width=width_frame, height=150)
lf_H1.grid_propagate(False)

label_month = tk.Label(lf_H1, text="Виберіть місяць: ", font=courier_14, foreground='red')
label_month.grid(row=0, column=0, ipadx=6, ipady=6, padx=5, pady=5, sticky=W)

selected_report = tk.StringVar(value=months[0])
report_menu = ttk.Combobox(lf_H1, textvariable=selected_report, values=months)
report_menu.tag = "month"
report_menu.grid(row=0, column=1,padx=5, pady=5, sticky=W)
report_menu.bind("<<ComboboxSelected>>", selected)

label_a = tk.Label(lf_H1, text="Виберіть адресу: ", font=courier_14, foreground='red')
label_a.grid(row=1, column=0, ipadx=6, ipady=6, padx=5, pady=5, sticky=W)

selected_a = tk.StringVar(value=address[0])
address_t = ttk.Combobox(lf_H1, textvariable=selected_a, values=address)
address_t.tag = 'address'
address_t.grid(row=1, column=1,padx=5, pady=5 )
address_t.bind("<<ComboboxSelected>>", selected)


# ========================  end main frame  =======================================

# ***************************** frame date count 8  **********************************

lf_H8 = ttk.Frame(lf_MF, borderwidth=10, relief=SUNKEN)
lf_H8.config(width=width_frame, height=300)
lf_H8.grid_propagate(False)

vcmd = (lf_H8.register(validate_input), "%P")


for i, d in enumerate(energy):  # створюємо 3 поля


    label = tk.Label(lf_H8, text=f"Введіть показники лічильника {d}: ",  font=courier_14, foreground='red')
    label.grid(row=i, column=0, ipadx=6, ipady=6, padx=5, pady=5)

    entry = tk.Entry(lf_H8, validate="key", validatecommand=vcmd)
    entry.grid(row=i, column=1, ipadx=6, ipady=6)

    entries.append(entry)

# ************************************ КНОПКА ДОБАВИТИ  ***********************************************************

save_btn = tk.Button(lf_H8, text="ЗБЕРЕГТИ ПОКАЗНИКИ", font=courier_10, state='normal', command=add_date)
save_btn.grid(row=3, column=0, ipadx=6, ipady=6, padx=50, pady=30)


lf_H1.grid(column=0, row=2, padx=20, pady=10, sticky=W)
lf_H8.grid(column=0, row=5, padx=20, pady=10, sticky=W)


lf_MF.grid(column=0, row=1, ipadx=6, ipady=6, padx=20, pady=20, sticky=W)

root.mainloop()