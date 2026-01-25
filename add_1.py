import tkinter as tk
from tkinter import *
from tkinter import font, messagebox
from tkinter import ttk

import openpyxl
from openpyxl import load_workbook

months = [" ", "Січень", "Лютий", "Березень", "Квітень", "Травень", "Червень", "Липень", "Серпень", "Вересень",
                  "Жовтень", "Листопад", "Грудень"]

energy = ['газу', 'електроенергії', 'води']

entries = []
adress = [' ', 'КВІТНЕВА-8', 'КВІТНЕВА-30-5']

text = ''

def selected(event):
    global text
    widget = event.widget
    tag = getattr(widget, "tag", None)
    text = widget.get()

    if tag == "month":
        label_report(lf_H1, 2, 0, text)



    if tag == "adress":
        label_report(lf_H1, 2, 1, text)




def add_date():

    '''
    if text == '':
        messagebox.showerror("Помилка вводу", "Оберите адресу !!!")
    elif text == "КВІТНЕВА-8":
        wb = load_workbook("K_8.xlsx")
        ws = wb.active
        data_gas_num = entries[0].get()
        data_light_num = entries[1].get()
        data_water_num = entries[2].get()

        ws['C58'] = data_gas_num
        ws['J58'] = data_light_num
        ws['P58'] = data_water_num
        wb.save("K_8.xlsx")


    elif text == "КВІТНЕВА-30-5":
        wb = load_workbook("k_30.xlsx")
        ws = wb.active
        data_gas_num = entries[0].get()
        data_light_num = entries[1].get()
        data_water_num = entries[2].get()

        ws['C45'] = data_gas_num
        ws['J45'] = data_light_num
        ws['P45'] = data_water_num
        wb.save("k_30.xlsx")


    else:
        print('error')
        messagebox.showerror("Помилка вводу", "Оберите адресу !!!") '''

    match text:
        case '':
            messagebox.showerror("Помилка вводу", "Оберите адресу !!!")
        case "КВІТНЕВА-8":
            wb = load_workbook("K_8.xlsx")
            ws = wb.active
            data_gas_num = entries[0].get()
            data_light_num = entries[1].get()
            data_water_num = entries[2].get()

            ws['C58'] = data_gas_num
            ws['J58'] = data_light_num
            ws['P58'] = data_water_num
            wb.save("K_8.xlsx")
        case "КВІТНЕВА-30-5":
            wb = load_workbook("k_30.xlsx")
            ws = wb.active
            data_gas_num = entries[0].get()
            data_light_num = entries[1].get()
            data_water_num = entries[2].get()

            ws['C45'] = data_gas_num
            ws['J45'] = data_light_num
            ws['P45'] = data_water_num
            wb.save("k_30.xlsx")








def label_report(frame, row, column, value):

    report_label = ttk.Label(frame, text=value, font=courier_14, foreground='green')
    report_label.grid(row=row, column=column, ipadx=6, ipady=6, padx=55, pady=5)
    print('l>>>>>')

def validate_input(new_value):
    if new_value == "":  # дозволяємо очищення поля
        return True
    if not new_value.isdigit():
        messagebox.showerror("Помилка вводу", "Дозволені тільки цифри!")
        return False
    if len(new_value) > 5:
        messagebox.showerror("Помилка вводу", "Максимум 5 цифр!")
        return False
    return True


# /////////////////////////////////  ///////////////////////////////////////////////////////////////////////////
root = tk.Tk()
root.title("MAIN WINDOW")
root.geometry("900x800")
root.resizable(False, False)

courier_10 = font.Font(family="Courier", size=10, weight=font.BOLD)
courier_14 = font.Font(family="Courier", size=14, weight=font.BOLD)
courier_18 = font.Font(family="Courier", size=18, weight=font.BOLD)
width_frame = 800

label = tk.Label(root, text="ДОДАВАННЯ ПОКАЗНИКІВ ЛІЧИЛЬНИКІВ У ФАЙЛ", fg="BLUE", font=courier_18)
label.grid(row=0, column=0, columnspan=3, ipadx=6, ipady=6, padx=5, pady=15)

# ========================  main frame  =======================================

lf_MF = ttk.Frame(root, borderwidth=10, relief=SUNKEN)
lf_MF.config(width=850, height=600)
# lf_MF.grid_propagate(False)

# label = tk.Label(lf_MF, text="ДОДАВАННЯ ПОКАЗНИКІВ КВІТНЕВА-8", fg="BLACK", font=courier_18)
# label.grid(row=0, column=0, columnspan=3, ipadx=6, ipady=6, padx=5, pady=15)

lf_H1 = ttk.Frame(lf_MF, borderwidth=10, relief=SUNKEN)
lf_H1.config(width=width_frame, height=150)
lf_H1.grid_propagate(False)

label_month = tk.Label(lf_H1, text="Виберіть місяць: ", font=courier_14, foreground='red')
label_month.grid(row=0, column=0, ipadx=6, ipady=6, padx=5, pady=5, sticky=W)

selected_report = tk.StringVar(value=months[0])
report_menu = ttk.Combobox(lf_H1, textvariable=selected_report, values=months)
report_menu.tag = "month"
report_menu.grid(row=0, column=1,padx=5, pady=5, sticky=W)
report_menu.bind("<<ComboboxSelected>>", selected)

label_a = tk.Label(lf_H1, text="Виберіть адресу: ", font=courier_14, foreground='red')
label_a.grid(row=1, column=0, ipadx=6, ipady=6, padx=5, pady=5, sticky=W)

selected_a = tk.StringVar(value=adress[0])
adress_t = ttk.Combobox(lf_H1, textvariable=selected_a, values=adress)
adress_t.tag = 'adress'
adress_t.grid(row=1, column=1,padx=5, pady=5 )
adress_t.bind("<<ComboboxSelected>>", selected)


# ========================  end main frame  =======================================

# ***************************** frame date count 8  **********************************

lf_H8 = ttk.Frame(lf_MF, borderwidth=10, relief=SUNKEN)
lf_H8.config(width=width_frame, height=300)
lf_H8.grid_propagate(False)

vcmd = (lf_H8.register(validate_input), "%P")

'''' # ************************************  ГАЗ  ****************************************************************

label_month = tk.Label(lf_H8, text="Введіть показники лічильника газу: ", font=courier_14, foreground='red')
label_month.grid(row=0, column=0, ipadx=6, ipady=6, padx=5, pady=5)

data_gas = ttk.Entry(lf_H8, validate="key", validatecommand=vcmd)
data_gas.grid(row=0, column=1, ipadx=6, ipady=6)




# *************************************  СВІТЛО  **********************************************************************

label_year = tk.Label(lf_H8, text="Введіть показники лічильника електроенергії: ", font=courier_14, foreground='red')
label_year.grid(row=1, column=0, ipadx=6, ipady=6, padx=5, pady=5)

data_light = ttk.Entry(lf_H8, validate="key", validatecommand=vcmd)
data_light.grid(row=1, column=1, ipadx=6, ipady=6)


# ************************************ ВОДА ***************************************************************************

label_year = tk.Label(lf_H8, text="Введіть показники лічильника води: ", font=courier_14, foreground='red')
label_year.grid(row=2, column=0, ipadx=6, ipady=6, padx=5, pady=5)

data_water = ttk.Entry(lf_H8, validate="key", validatecommand=vcmd)
data_water.grid(row=2, column=1, ipadx=6, ipady=6)

'''
for i, d in enumerate(energy):  # створюємо 3 поля


    label = tk.Label(lf_H8, text=f"Введіть показники лічильника {d}: ",  font=courier_14, foreground='red')
    label.grid(row=i, column=0, ipadx=6, ipady=6, padx=5, pady=5)

    entry = tk.Entry(lf_H8, validate="key", validatecommand=vcmd)
    entry.grid(row=i, column=1, ipadx=6, ipady=6)

    entries.append(entry)

# ************************************ КНОПКА ДОБАВИТИ  ***********************************************************

save_btn = tk.Button(lf_H8, text="ЗБЕРЕГТИ ПОКАЗНИКИ", font=courier_10, state='normal', command=add_date)
save_btn.grid(row=3, column=0, ipadx=6, ipady=6, padx=50, pady=30)


lf_H1.grid(column=0, row=2, padx=20, pady=10, sticky=W)
lf_H8.grid(column=0, row=5, padx=20, pady=10, sticky=W)


lf_MF.grid(column=0, row=1, ipadx=6, ipady=6, padx=20, pady=20, sticky=W)

root.mainloop()

'''
from openpyxl import load_workbook

wb = load_workbook("example.xlsx")
ws = wb.active

ws['C3'] = "Нові дані"
wb.save("example.xlsx")


C, J, P 
'''